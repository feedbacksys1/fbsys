"""
Представления основного приложения.
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.http import HttpResponse, FileResponse, Http404
from django.contrib.auth import login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import get_user_model
from django.db.models import Q
from django.utils import timezone

from django.contrib import messages

from .forms import (
    RegistrationForm, LoginForm, ProfileSettingsForm, PasswordChangeFormStyled,
    StudentRequestForm, RequestReplyForm, CuratorRequestTypeForm,
    REQUEST_ATTACHMENT_EXTENSIONS, REQUEST_ATTACHMENT_MAX_SIZE,
)
from .models import Profile, Role, GeneralFeedback, FeedbackStatus, StudentRequest, StudentRequestMessage, RequestAttachment

User = get_user_model()


def landing(request):
    """
    Главная (лендинг) страница системы обслуживания заявок.
    """
    return render(request, 'landing.html')


def contacts(request):
    """
    Страница контактов.
    """
    return render(request, 'contacts.html')


def about(request):
    """
    Статичная страница «О нас».
    """
    return render(request, 'about.html')


def advantages(request):
    """
    Статичная страница «Преимущества».
    """
    return render(request, 'advantages.html')


def request_rules(request):
    """
    Статичная страница «Правила подачи заявок».
    """
    return render(request, 'request_rules.html')


def student_guide(request):
    """
    Статичная страница «Инструкция для студентов».
    """
    return render(request, 'student_guide.html')


def faq(request):
    """
    Статичная страница «Вопросы и ответы».
    """
    return render(request, 'faq.html')


def feedback(request):
    """
    Страница обратной связи (общая форма для посетителей сайта).
    """
    if request.method == 'POST':
        name = (request.POST.get('name') or '').strip()
        email = (request.POST.get('email') or '').strip()
        phone = (request.POST.get('phone') or '').strip()
        topic = (request.POST.get('topic') or '').strip()
        message_text = (request.POST.get('message') or '').strip()
        if not name or not email or not message_text:
            messages.error(request, 'Заполните обязательные поля: Имя, Email, Сообщение.')
            return render(request, 'feedback.html', {
                'form_name': name,
                'form_email': email,
                'form_phone': phone,
                'form_topic': topic,
                'form_message': message_text,
            })
        try:
            GeneralFeedback.objects.create(
                name=name,
                email=email,
                phone=phone,
                topic=topic,
                message=message_text,
            )
            messages.success(request, 'Ваше сообщение отправлено. Мы рассмотрим его в ближайшее время.')
            return redirect('feedback')
        except Exception:
            messages.error(request, 'Не удалось отправить сообщение. Попробуйте позже.')
            return render(request, 'feedback.html', {
                'form_name': name,
                'form_email': email,
                'form_phone': phone,
                'form_topic': topic,
                'form_message': message_text,
            })
    return render(request, 'feedback.html')


def login_view(request):
    """
    Страница входа.
    """
    if request.user.is_authenticated:
        return redirect('landing')
    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('landing')
    else:
        form = LoginForm(request)
    return render(request, 'login.html', {'form': form})


def register(request):
    """
    Страница регистрации.
    """
    if request.user.is_authenticated:
        return redirect('landing')
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('landing')
    else:
        form = RegistrationForm()
    return render(request, 'register.html', {'form': form})


def logout_view(request):
    """
    Выход из учётной записи (GET для перехода по ссылке).
    """
    logout(request)
    return redirect('landing')


@login_required
def settings_view(request):
    """
    Настройки аккаунта: данные профиля и смена пароля.
    """
    user = request.user
    try:
        profile = user.profile
    except Profile.DoesNotExist:
        profile = None
    profile_form = ProfileSettingsForm(
        request.POST if request.method == 'POST' and 'profile_submit' in request.POST else None,
        user=user,
        initial={
            'first_name': user.first_name,
            'last_name': user.last_name,
            'email': user.email or '',
            'username': user.username,
            'student_number': profile.student_number if profile else '',
        },
    )
    password_form = PasswordChangeFormStyled(
        user,
        request.POST if request.method == 'POST' and 'password_submit' in request.POST else None,
    )
    if request.method == 'POST':
        if 'profile_submit' in request.POST and profile_form.is_valid():
            user.first_name = profile_form.cleaned_data['first_name']
            user.last_name = profile_form.cleaned_data['last_name']
            user.email = profile_form.cleaned_data['email']
            user.username = profile_form.cleaned_data['username']
            user.save()
            prof, _ = Profile.objects.get_or_create(user=user, defaults={'role': Role.STUDENT})
            prof.student_number = profile_form.cleaned_data.get('student_number', '')
            prof.save()
            messages.success(request, 'Данные аккаунта сохранены.')
            return redirect('settings')
        if 'password_submit' in request.POST and password_form.is_valid():
            password_form.save()
            update_session_auth_hash(request, password_form.user)
            messages.success(request, 'Пароль успешно изменён.')
            return redirect('settings')
    return render(request, 'settings.html', {
        'profile_form': profile_form,
        'password_form': password_form,
    })


def _is_staff(user):
    return user.is_authenticated and user.is_staff


def _validate_attachment(file):
    """Проверка расширения и размера файла. Возвращает (ok, error_message)."""
    ext = (file.name.split('.')[-1] or '').lower()
    if ext not in REQUEST_ATTACHMENT_EXTENSIONS:
        return False, f'Недопустимый тип файла. Разрешены: {", ".join(sorted(REQUEST_ATTACHMENT_EXTENSIONS))}'
    if file.size > REQUEST_ATTACHMENT_MAX_SIZE:
        return False, 'Размер файла не должен превышать 10 МБ.'
    return True, None


@login_required
def cabinet_view(request):
    """
    Личный кабинет для студентов и панель управления для кураторов.
    Переписка по заявке, вложения к заявке.
    """
    if request.user.is_staff:
        return redirect('admin_panel')
    try:
        profile = request.user.profile
        role = profile.role
    except Profile.DoesNotExist:
        role = Role.STUDENT
    page_title = 'Панель управления' if role == Role.TEACHER else 'Личный кабинет'

    request_form = None
    detail_request = None
    thread_messages = []
    reply_form = None
    redirect_detail = None  # для редиректа с сохранением ?detail=id
    curator_request_type_form = None

    # --- POST: сохранение типа запроса (куратор) ---
    if role == Role.TEACHER and request.method == 'POST' and 'save_request_type' in request.POST:
        curator_request_type_form = CuratorRequestTypeForm(request.POST)
        if curator_request_type_form.is_valid():
            prof, _ = Profile.objects.get_or_create(user=request.user, defaults={'role': Role.TEACHER})
            prof.request_type = (curator_request_type_form.cleaned_data.get('request_type') or '').strip()[:200]
            prof.save(update_fields=['request_type'])
            messages.success(request, 'Тип запроса сохранён.')
            return redirect('cabinet')

    # --- POST: ответ в переписке (куратор или студент) ---
    if request.method == 'POST' and request.POST.get('reply_body') is not None:
        reply_form = RequestReplyForm({'body': request.POST.get('reply_body', '')})
        if reply_form.is_valid():
            req_id = request.POST.get('request_id')
            try:
                req_id = int(req_id)
                req = StudentRequest.objects.get(pk=req_id)
            except (TypeError, ValueError, StudentRequest.DoesNotExist):
                req = None
            if req and role == Role.TEACHER and req.recipient_id == request.user.pk:
                StudentRequestMessage.objects.create(
                    request=req,
                    author=request.user,
                    body=reply_form.cleaned_data['body'].strip(),
                )
                if not req.thread_opened_at:
                    req.thread_opened_at = timezone.now()
                    req.save(update_fields=['thread_opened_at'])
                messages.success(request, 'Ответ отправлен.')
                qs = '?detail=' + str(req_id)
                if request.GET.get('status'):
                    qs += '&status=' + request.GET.get('status')
                return redirect(reverse('cabinet') + qs)
            elif req and role == Role.STUDENT and req.sender_id == request.user.pk and req.thread_opened_at:
                StudentRequestMessage.objects.create(
                    request=req,
                    author=request.user,
                    body=reply_form.cleaned_data['body'].strip(),
                )
                messages.success(request, 'Ответ отправлен.')
                qs = '?detail=' + str(req_id)
                if request.GET.get('status'):
                    qs += '&status=' + request.GET.get('status')
                return redirect(reverse('cabinet') + qs)
            else:
                messages.error(request, 'Не удалось отправить ответ.')
        else:
            redirect_detail = request.POST.get('request_id')

    # --- POST: смена статуса (куратор) ---
    if request.method == 'POST' and request.POST.get('new_status') and not redirect_detail:
        request_id = request.POST.get('request_id')
        new_status = request.POST.get('new_status')
        if request_id and new_status and new_status in dict(FeedbackStatus.choices):
            try:
                req = StudentRequest.objects.get(pk=int(request_id), recipient=request.user)
                req.status = new_status
                req.status_changed_at = timezone.now()
                req.save()
                messages.success(request, 'Статус заявки обновлён.')
                redirect_status = request.POST.get('redirect_status')
                redirect_detail = int(request_id)
                if redirect_status and redirect_status in dict(FeedbackStatus.choices):
                    return redirect(reverse('cabinet') + '?detail=' + str(redirect_detail) + '&status=' + redirect_status)
                return redirect(reverse('cabinet') + '?detail=' + str(redirect_detail))
            except (StudentRequest.DoesNotExist, ValueError):
                pass

    # --- POST: новая заявка (студент) с вложениями ---
    if role == Role.STUDENT and request.method == 'POST' and not request.POST.get('request_id'):
        request_form = StudentRequestForm(request.POST)
        if request_form.is_valid():
            req = StudentRequest.objects.create(
                sender=request.user,
                recipient=request_form.cleaned_data['recipient'],
                topic=(request_form.cleaned_data.get('topic') or '').strip(),
                message=request_form.cleaned_data['message'].strip(),
            )
            files = request.FILES.getlist('attachments')
            for f in files:
                ok, err = _validate_attachment(f)
                if not ok:
                    messages.warning(request, err)
                    continue
                original_name = f.name
                RequestAttachment.objects.create(request=req, file=f, original_name=original_name)
            messages.success(request, 'Заявка успешно отправлена.')
            return redirect('cabinet')

    if role == Role.STUDENT and not request_form:
        request_form = StudentRequestForm(request.POST if request.method == 'POST' else None)

    # --- Загрузка заявки для детального просмотра (куратор или студент) ---
    detail_id = request.GET.get('detail') or (redirect_detail and str(redirect_detail))
    if detail_id:
        try:
            req = StudentRequest.objects.get(pk=int(detail_id))
            if role == Role.TEACHER and req.recipient_id == request.user.pk:
                detail_request = req
            elif role == Role.STUDENT and req.sender_id == request.user.pk:
                detail_request = req
        except (StudentRequest.DoesNotExist, ValueError):
            pass
        if detail_request:
            thread_messages = list(detail_request.thread_messages.select_related('author').order_by('created_at'))
            reply_form = reply_form or RequestReplyForm()

    filter_status = request.GET.get('status')
    if filter_status not in dict(FeedbackStatus.choices):
        filter_status = None

    context = {
        'page_title': page_title,
        'role': role,
        'request_form': request_form,
        'detail_request': detail_request,
        'thread_messages': thread_messages,
        'reply_form': reply_form,
        'status_choices': list(FeedbackStatus.choices),
        'filter_status': filter_status,
        'attachment_extensions': ', '.join(sorted(REQUEST_ATTACHMENT_EXTENSIONS)),
    }
    if role == Role.STUDENT:
        qs = StudentRequest.objects.filter(sender=request.user).select_related('recipient')
        if filter_status:
            qs = qs.filter(status=filter_status)
        context['sent_requests'] = qs
    if role == Role.TEACHER:
        qs = StudentRequest.objects.filter(recipient=request.user).select_related('sender')
        if filter_status:
            qs = qs.filter(status=filter_status)
        context['received_requests'] = qs
        try:
            profile = request.user.profile
            initial_type = getattr(profile, 'request_type', '') or ''
        except Profile.DoesNotExist:
            initial_type = ''
        context['curator_request_type_form'] = curator_request_type_form or CuratorRequestTypeForm(initial={'request_type': initial_type})
    return render(request, 'cabinet.html', context)


@login_required
def download_request_attachment(request, attachment_id):
    """
    Скачивание вложения к заявке. Доступно куратору (получатель заявки) и студенту (отправитель заявки).
    """
    attachment = get_object_or_404(RequestAttachment, pk=attachment_id)
    req = attachment.request
    # Куратор — получатель заявки; студент — отправитель
    if request.user.pk != req.recipient_id and request.user.pk != req.sender_id:
        raise Http404
    try:
        file_handle = attachment.file.open('rb')
        response = FileResponse(file_handle, as_attachment=True, filename=attachment.original_name)
        return response
    except (OSError, ValueError):
        raise Http404


@login_required
@user_passes_test(_is_staff, login_url='landing')
def admin_panel(request):
    """
    Редирект на страницу «Просмотр статистики» (стартовая страница админ-панели).
    """
    return redirect('admin_statistics')


@login_required
@user_passes_test(_is_staff, login_url='landing')
def admin_statistics(request):
    """
    Просмотр статистики по заявкам по каждому куратору (только для staff).
    """
    search_q = request.GET.get('q', '').strip()
    stats = _get_teacher_stats(search_q)
    return render(request, 'admin_statistics.html', {'stats': stats, 'search_q': search_q})


def _get_teacher_stats(search_q=None):
    """Возвращает список словарей со статистикой по кураторам (для страницы и экспорта)."""
    teachers = User.objects.filter(profile__role=Role.TEACHER).order_by('last_name', 'first_name')
    if search_q:
        teachers = teachers.filter(
            Q(username__icontains=search_q)
            | Q(first_name__icontains=search_q)
            | Q(last_name__icontains=search_q)
            | Q(email__icontains=search_q),
        )
    stats = []
    for t in teachers:
        total = StudentRequest.objects.filter(recipient=t).count()
        reviewed = StudentRequest.objects.filter(recipient=t, status=FeedbackStatus.REVIEWED).count()
        awaiting = StudentRequest.objects.filter(
            recipient=t,
            status__in=[FeedbackStatus.NEW, FeedbackStatus.IN_PROGRESS],
        ).count()
        stats.append({
            'teacher': t,
            'total': total,
            'reviewed': reviewed,
            'awaiting': awaiting,
        })
    return stats


@login_required
@user_passes_test(_is_staff, login_url='landing')
def admin_statistics_export(request):
    """
    Скачивание статистики по кураторам в виде Excel (только для staff).
    Учитывается текущий поисковый запрос (параметр q).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from io import BytesIO

    search_q = request.GET.get('q', '').strip()
    stats = _get_teacher_stats(search_q)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Статистика'

    headers = ['Куратор', 'Всего заявок', 'Рассмотрено', 'Ожидают рассмотрения']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for row, item in enumerate(stats, 2):
        teacher_name = item['teacher'].get_full_name() or item['teacher'].get_username()
        ws.cell(row=row, column=1, value=teacher_name)
        ws.cell(row=row, column=2, value=item['total'])
        ws.cell(row=row, column=3, value=item['reviewed'])
        ws.cell(row=row, column=4, value=item['awaiting'])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="statistics.xlsx"'
    return response


@login_required
@user_passes_test(_is_staff, login_url='landing')
def admin_users(request):
    """
    Управление пользователями: список и смена роли (только для staff).
    """
    if request.method == 'POST':
        for key, new_role in request.POST.items():
            if key.startswith('role_') and new_role in dict(Role.choices):
                try:
                    user_id = int(key.replace('role_', ''))
                except ValueError:
                    continue
                user = get_object_or_404(User, pk=user_id)
                profile, _ = Profile.objects.get_or_create(user=user, defaults={'role': Role.STUDENT})
                profile.role = new_role
                profile.save()
                user.is_staff = new_role == Role.ADMIN
                user.save(update_fields=['is_staff'])
        from django.urls import reverse
        from urllib.parse import urlencode
        url = reverse('admin_users')
        q = request.POST.get('q', '').strip()
        if q:
            url += '?' + urlencode({'q': q})
        return redirect(url)
    users = User.objects.all().order_by('id')
    search_q = request.GET.get('q', '').strip()
    if search_q:
        users = users.filter(
            Q(username__icontains=search_q) | Q(email__icontains=search_q)
        )
    users_with_profiles = []
    for u in users:
        try:
            profile = u.profile
        except Profile.DoesNotExist:
            profile = None
        users_with_profiles.append({'user': u, 'profile': profile})
    return render(request, 'admin_users.html', {
        'users_with_profiles': users_with_profiles,
        'role_choices': Role.choices,
        'search_q': search_q,
    })


@login_required
@user_passes_test(_is_staff, login_url='landing')
def admin_feedback_list(request):
    """
    Просмотр обращений с сайта и смена статусов (только для staff).
    """
    if request.method == 'POST':
        now = timezone.now()
        for key, new_status in request.POST.items():
            if key.startswith('status_') and new_status in dict(FeedbackStatus.choices):
                try:
                    fb_id = int(key.replace('status_', ''))
                except ValueError:
                    continue
                fb = get_object_or_404(GeneralFeedback, pk=fb_id)
                if fb.status != new_status:
                    fb.status = new_status
                    fb.status_changed_at = now
                    fb.status_changed_by = request.user
                    fb.save()
        from django.urls import reverse
        from urllib.parse import urlencode
        url = reverse('admin_feedback_list')
        status_param = request.POST.get('filter_status', '').strip()
        if status_param and status_param in dict(FeedbackStatus.choices):
            url += '?' + urlencode({'status': status_param})
        return redirect(url)
    feedback_list = GeneralFeedback.objects.all()
    filter_status = request.GET.get('status', '').strip()
    if filter_status and filter_status in dict(FeedbackStatus.choices):
        feedback_list = feedback_list.filter(status=filter_status)
    return render(request, 'admin_feedback_list.html', {
        'feedback_list': feedback_list,
        'status_choices': FeedbackStatus.choices,
        'filter_status': filter_status,
    })
